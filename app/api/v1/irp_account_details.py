from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal, InvalidOperation
from app.database import get_db
from app.schemas.irp_account_detail import IRPAccountDetail, IRPAccountDetailCreate, IRPAccountDetailUpdate
from app.crud.irp_account_detail import (
    get_irp_account_details as crud_get_all,
    get_irp_account_detail as crud_get_one,
    create_irp_account_detail as crud_create,
    update_irp_account_detail as crud_update,
    delete_irp_account_detail as crud_delete,
    get_irp_account_detail_by_stock_code
)

router = APIRouter()

@router.get("/account/{account_id}", response_model=List[IRPAccountDetail])
def read_irp_account_details(
    account_id: int,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    try:
        details = crud_get_all(db, account_id=account_id, skip=skip, limit=limit)
        return details
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"서버 오류: {str(e)}")

@router.get("/{detail_id}", response_model=IRPAccountDetail)
def read_irp_account_detail(detail_id: int, db: Session = Depends(get_db)):
    db_detail = crud_get_one(db, detail_id=detail_id)
    if db_detail is None:
        raise HTTPException(status_code=404, detail="IRP 계좌 상세를 찾을 수 없습니다")
    return db_detail

@router.post("/", response_model=IRPAccountDetail, status_code=201)
def create_irp_account_detail(detail: IRPAccountDetailCreate, db: Session = Depends(get_db)):
    try:
        return crud_create(db=db, detail=detail)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/{detail_id}", response_model=IRPAccountDetail)
def update_irp_account_detail(
    detail_id: int,
    detail: IRPAccountDetailUpdate,
    db: Session = Depends(get_db)
):
    try:
        return crud_update(db=db, detail_id=detail_id, detail=detail)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{detail_id}", status_code=204)
def delete_irp_account_detail(detail_id: int, db: Session = Depends(get_db)):
    try:
        crud_delete(db=db, detail_id=detail_id)
        return None
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/template/download")
def download_template():
    """IRP 종목 상세 엑셀 템플릿 다운로드"""
    wb = Workbook()
    ws = wb.active
    ws.title = "IRP 종목 상세"
    
    # 헤더 설정
    headers = ["종목코드", "수량", "매입단가", "현재가", "매수수수료"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 15  # 종목코드
    ws.column_dimensions['B'].width = 15  # 수량
    ws.column_dimensions['C'].width = 15  # 매입단가
    ws.column_dimensions['D'].width = 15  # 현재가
    ws.column_dimensions['E'].width = 15  # 매수수수료
    
    # 예시 데이터 행 추가
    example_row = ["069500", "10", "50000", "52000", "0"]
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=irp_account_detail_template.xlsx"}
    )

@router.post("/upload/{account_id}")
def upload_excel(
    account_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """IRP 종목 상세 엑셀 파일 업로드"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")
    
    try:
        # 파일 읽기
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        errors = []
        success_count = 0
        update_count = 0
        create_count = 0
        
        # 첫 번째 행은 헤더이므로 건너뛰기
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not any(row):
                continue
            
            try:
                stock_code = str(row[0]).strip() if row[0] else None
                quantity = row[1]
                purchase_avg_price = row[2]
                current_price = row[3]
                purchase_fee = row[4] if len(row) > 4 else 0
                
                # 필수 필드 검증
                if not stock_code:
                    errors.append(f"{row_idx}행: 종목코드가 비어있습니다.")
                    continue
                
                # 종목코드 6자리 검증 (문자 포함 가능)
                if len(stock_code) != 6:
                    errors.append(f"{row_idx}행: 종목코드는 6자리여야 합니다. (현재: {len(stock_code)}자리)")
                    continue
                
                # 숫자 검증
                try:
                    quantity = Decimal(str(quantity)) if quantity is not None else None
                    purchase_avg_price = Decimal(str(purchase_avg_price)) if purchase_avg_price is not None else None
                    current_price = Decimal(str(current_price)) if current_price is not None else None
                    purchase_fee = Decimal(str(int(float(purchase_fee)))) if purchase_fee is not None else Decimal(0)
                except (ValueError, TypeError, InvalidOperation) as e:
                    errors.append(f"{row_idx}행: 숫자가 아닌 값이 있습니다. (수량, 매입단가, 현재가, 매수수수료는 숫자여야 합니다.)")
                    continue
                
                if quantity is None or purchase_avg_price is None or current_price is None:
                    errors.append(f"{row_idx}행: 필수 필드(수량, 매입단가, 현재가)가 비어있습니다.")
                    continue
                
                # 음수 검증
                if quantity < 0 or purchase_avg_price < 0 or current_price < 0 or purchase_fee < 0:
                    errors.append(f"{row_idx}행: 음수는 입력할 수 없습니다.")
                    continue
                
                # 기존 레코드 확인
                existing_detail = get_irp_account_detail_by_stock_code(db, account_id, stock_code)
                
                if existing_detail:
                    # 기존 레코드 업데이트
                    detail_data = IRPAccountDetailUpdate(
                        account_id=account_id,
                        stock_code=stock_code,
                        quantity=quantity,
                        purchase_avg_price=purchase_avg_price,
                        current_price=current_price,
                        purchase_fee=purchase_fee,
                        sale_fee=existing_detail.sale_fee  # 기존 매도수수료 유지
                    )
                    crud_update(db=db, detail_id=existing_detail.id, detail=detail_data)
                    update_count += 1
                else:
                    # 새 레코드 생성
                    detail_data = IRPAccountDetailCreate(
                        account_id=account_id,
                        stock_code=stock_code,
                        quantity=quantity,
                        purchase_avg_price=purchase_avg_price,
                        current_price=current_price,
                        purchase_fee=purchase_fee,
                        sale_fee=Decimal(0)  # 매도수수료는 기본값 0
                    )
                    crud_create(db=db, detail=detail_data)
                    create_count += 1
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"{row_idx}행: 처리 중 오류 발생 - {str(e)}")
        
        db.commit()
        
        result = {
            "success_count": success_count,
            "create_count": create_count,
            "update_count": update_count,
            "error_count": len(errors),
            "errors": errors
        }
        
        if errors:
            raise HTTPException(status_code=400, detail=result)
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 처리 중 오류 발생: {str(e)}")

